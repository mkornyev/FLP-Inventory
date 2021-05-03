#!/usr/bin/python
from django.core.management.base import BaseCommand
from inventory.models import Item, Category
import openpyxl

class Command(BaseCommand):
    args = 'this function takes two arguments, the paths to the \
            Sharepoint manage inventory excel file, Sharepoint manage items excel file'
    help = 'Download the "Manage Inventory" and "Manage Items"  pages\' from Sharepoint \
            and provide the path to these files as arguments'
    
    def add_arguments(self, parser):
        parser.add_argument('manage_inventory')
        parser.add_argument('manage_items')

    def handle(self, *args, **options):
        INVENTORY_FILE_PATH = options['manage_inventory']
        ITEMS_FILE_PATH = options['manage_items']

        inventory_sheet = openpyxl.load_workbook(INVENTORY_FILE_PATH).active
        items_sheet = openpyxl.load_workbook(ITEMS_FILE_PATH).active

        seen_items = set()
        seen_categories = set()

        for r in items_sheet.iter_rows(min_row=2):
            item_name = r[1].value
            item_category = r[2].value
            print(f"adding {item_name} w category {item_category}")
            qs = Category.objects.filter(name__exact="Clothing")
            if len(qs) == 0:
                if item_category not in seen_items:
                    Category.objects.create(name=item_category)
                    seen_categories.add(item_category)
            elif item_name not in seen_items:
                Item.objects.create(category=qs[0], name=item_name, quantity=0)
                seen_items.add(item_name)

        for r in inventory_sheet.iter_rows(min_row=2): # skip header
            item_name = r[1].value
            quantity = r[4].value
            print(f"updating quantity {item_name} w {quantity}")
            qs = Item.objects.filter(name__exact=item_name)
            if len(qs) != 0:
                qs[0].quantity = quantity
                qs[0].save()
            else:
                print(f"{item_name} not found")
            
